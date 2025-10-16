import warnings
warnings.filterwarnings('ignore')
import pandas as pd
import os
import sqlite3
import tempfile
import time
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage, SystemMessage
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import win32com.client as win32
import pythoncom
from fpdf import FPDF

# API配置
API_KEY = "sk-062b1fc6f6a9425eae59020faa7b5f6e"
API_BASE = "https://api.deepseek.com/v1"
# 模型配置
MODEL_NAME = "deepseek-chat"
MODEL_TEMPERATURE = 0

# 输出路径（可根据需要修改）
OUTPUT_PATH = "F:\\analysis\\data"


class DataProcessor:
    def __init__(self):
        # 初始化大模型
        self.llm = ChatOpenAI(
            openai_api_key=API_KEY,
            openai_api_base=API_BASE,
            model_name=MODEL_NAME,
            temperature=MODEL_TEMPERATURE
        )
        # 确保输出目录存在
        self._ensure_output_directory()
        # 存储数据库连接引用
        self.raw_connection = None
        self.current_table_name = None

    def _ensure_output_directory(self):
        """确保输出目录存在"""
        try:
            os.makedirs(OUTPUT_PATH, exist_ok=True)
            print(f"输出目录已确认: {OUTPUT_PATH}")
        except Exception as e:
            print(f"创建输出目录失败: {str(e)}")
            raise

    def create_in_memory_db(self, file_path, sample_ratio=1.0):
        """创建内存数据库并导入数据"""
        try:
            # 读取数据文件
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            elif file_path.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_path)
            else:
                return None, "错误：仅支持CSV和Excel格式"

            print(f"原始数据集大小: {len(df)}行, {len(df.columns)}列")

            # 大型数据集抽样处理
            if len(df) > 10000 and sample_ratio < 1.0:
                print(f"检测到大型数据集，使用{int(sample_ratio*100)}%抽样数据加速处理")
                df = df.sample(frac=sample_ratio, random_state=42)
                print(f"抽样后数据集大小: {len(df)}行")

            # 创建内存数据库
            conn = sqlite3.connect(":memory:")
            self.raw_connection = conn
            
            # 优化数据库写入性能
            conn.execute("PRAGMA synchronous = OFF")
            conn.execute("PRAGMA journal_mode = MEMORY")
            
            # 生成安全的表名
            table_name = "data_table"
            self.current_table_name = table_name
            
            # 写入数据到内存表
            df.to_sql(table_name, conn, if_exists='replace', index=False)
            
            print(f"数据已导入内存数据库，表名: {table_name}")
            print(f"列信息: {', '.join(df.columns.tolist())}")
            
            return df, table_name

        except Exception as e:
            self._cleanup_connections()
            return None, f"创建内存数据库失败: {str(e)}"

    def get_basic_info(self, table_name):
        """获取基本信息"""
        try:
            # 行数
            row_count_query = f"SELECT COUNT(*) as 行数 FROM {table_name}"
            row_count = pd.read_sql_query(row_count_query, self.raw_connection)
            
            # 列信息
            column_info = pd.read_sql_query(f"PRAGMA table_info({table_name})", self.raw_connection)
            
            result = f"表名: {table_name}\n"
            result += f"行数: {row_count.iloc[0, 0]}\n"
            result += f"列数: {len(column_info)}\n"
            result += "列信息:\n"
            
            for _, row in column_info.iterrows():
                result += f"  - {row['name']} ({row['type']})\n"
                
            return result
        except Exception as e:
            return f"获取基本信息失败: {str(e)}"

    def get_missing_value_analysis(self, table_name):
        """缺失值分析 - 修复版本"""
        try:
            column_info = pd.read_sql_query(f"PRAGMA table_info({table_name})", self.raw_connection)
            result = "缺失值分析:\n"
            
            total_count_query = f"SELECT COUNT(*) as total FROM {table_name}"
            total_count = pd.read_sql_query(total_count_query, self.raw_connection)
            total_rows = total_count.iloc[0, 0]
            
            for _, row in column_info.iterrows():
                col_name = row['name']
                # 检查该列的缺失值数量
                missing_query = f"SELECT COUNT(*) as missing FROM {table_name} WHERE `{col_name}` IS NULL"
                missing_count = pd.read_sql_query(missing_query, self.raw_connection)
                missing_num = missing_count.iloc[0, 0]
                missing_ratio = (missing_num / total_rows) * 100 if total_rows > 0 else 0
                result += f"  - {col_name}: {missing_num} 个缺失值 ({missing_ratio:.2f}%)\n"
                
            return result
        except Exception as e:
            return f"缺失值分析失败: {str(e)}"

    def get_duplicate_analysis(self, table_name):
        """重复数据分析 - 修复版本"""
        try:
            column_info = pd.read_sql_query(f"PRAGMA table_info({table_name})", self.raw_connection)
            
            # 获取所有列名
            columns = [row['name'] for _, row in column_info.iterrows()]
            columns_str = ", ".join([f"`{col}`" for col in columns])
            
            # 检查完全重复的行
            duplicate_query = f"""
            SELECT COUNT(*) as duplicate_count 
            FROM (
                SELECT {columns_str}, COUNT(*) as cnt 
                FROM {table_name} 
                GROUP BY {columns_str} 
                HAVING cnt > 1
            )
            """
            
            duplicate_result = pd.read_sql_query(duplicate_query, self.raw_connection)
            duplicate_count = duplicate_result.iloc[0, 0]
            
            result = "重复数据分析:\n"
            result += f"完全重复的行数: {duplicate_count}\n"
            
            # 检查关键字段的重复（比如序号）
            if any('序号' in col for col in columns):
                id_col = [col for col in columns if '序号' in col][0]
                id_duplicate_query = f"""
                SELECT COUNT(*) as duplicate_ids 
                FROM (
                    SELECT `{id_col}`, COUNT(*) as cnt 
                    FROM {table_name} 
                    GROUP BY `{id_col}` 
                    HAVING cnt > 1
                )
                """
                id_duplicate_result = pd.read_sql_query(id_duplicate_query, self.raw_connection)
                duplicate_ids = id_duplicate_result.iloc[0, 0]
                result += f"序号重复的数量: {duplicate_ids}\n"
                
            return result
        except Exception as e:
            return f"重复数据分析失败: {str(e)}"

    def get_numeric_analysis(self, table_name):
        """数值列分析 - 修复版本"""
        try:
            column_info = pd.read_sql_query(f"PRAGMA table_info({table_name})", self.raw_connection)
            result = "数值列分析:\n"
            
            # 识别数值列
            numeric_columns = []
            for _, row in column_info.iterrows():
                col_type = row['type'].upper()
                col_name = row['name']
                if 'INT' in col_type or 'REAL' in col_type or 'FLOAT' in col_type or 'NUM' in col_type:
                    numeric_columns.append(col_name)
                # 如果列名包含数字相关的词，也认为是数值列
                elif any(keyword in col_name.lower() for keyword in ['宽度', '高度', '坐标', '序号', '数量', '值']):
                    numeric_columns.append(col_name)
            
            for col in numeric_columns:
                try:
                    stats_query = f"""
                    SELECT 
                        COUNT(`{col}`) as 非空数量,
                        MIN(`{col}`) as 最小值, 
                        MAX(`{col}`) as 最大值, 
                        AVG(`{col}`) as 平均值,
                        SUM(CASE WHEN `{col}` IS NULL THEN 1 ELSE 0 END) as 空值数量
                    FROM {table_name}
                    """
                    stats = pd.read_sql_query(stats_query, self.raw_connection)
                    if hasattr(stats, 'iloc'):
                        count_val = stats.iloc[0, 0]
                        min_val = stats.iloc[0, 1]
                        max_val = stats.iloc[0, 2]
                        avg_val = stats.iloc[0, 3]
                        null_count = stats.iloc[0, 4]
                        result += f"  - {col}: 值范围[{min_val}, {max_val}], 平均值={avg_val:.2f}, 非空值={count_val}, 空值={null_count}\n"
                except Exception as col_error:
                    result += f"  - {col}: 分析出错 - {str(col_error)}\n"
                    
            return result
        except Exception as e:
            return f"数值列分析失败: {str(e)}"

    def get_text_analysis(self, table_name):
        """文本列分析 - 修复版本"""
        try:
            column_info = pd.read_sql_query(f"PRAGMA table_info({table_name})", self.raw_connection)
            result = "文本列分析:\n"
            
            # 识别文本列和可能的文本列
            text_columns = []
            
            for _, row in column_info.iterrows():
                col_type = row['type'].upper()
                col_name = row['name']
                if 'TEXT' in col_type or 'CHAR' in col_type:
                    text_columns.append(col_name)
            
            if not text_columns:
                result += "  未发现文本列，所有列均为数值类型\n"
                return result
            
            # 分析文本列
            for col in text_columns:
                try:
                    # 唯一值数量
                    distinct_query = f"SELECT COUNT(DISTINCT `{col}`) as distinct_count FROM {table_name}"
                    distinct_result = pd.read_sql_query(distinct_query, self.raw_connection)
                    distinct_count = distinct_result.iloc[0, 0]
                    
                    # 样本值
                    sample_query = f"SELECT `{col}` FROM {table_name} WHERE `{col}` IS NOT NULL LIMIT 5"
                    sample_result = pd.read_sql_query(sample_query, self.raw_connection)
                    sample_values = sample_result[col].tolist()
                    
                    result += f"  - {col}: {distinct_count} 个唯一值, 样本: {sample_values}\n"
                except Exception as col_error:
                    result += f"  - {col}: 分析出错\n"
            
            return result
        except Exception as e:
            return f"文本列分析失败: {str(e)}"

    def _create_temp_excel(self, df, max_rows=200):
        """创建临时Excel（优化格式）"""
        try:
            # 限制最大行数，避免生成过大文件
            df = df.head(max_rows)
            
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
                temp_path = temp_excel.name

            wb = Workbook()
            ws = wb.active
            ws.title = "清洗后数据"

            # 写入数据并设置格式
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    value = str(cell_value) if pd.notna(cell_value) else ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 文字自动换行
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical='center',
                        horizontal='left'
                    )

            # 自动调整列宽
            self._auto_adjust_column_width(ws, df)

            # 设置行高
            ws.row_dimensions[1].height = 28
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 35

            # 添加边框
            self._add_cell_borders(ws)

            wb.save(temp_path)
            return temp_path
        except Exception as e:
            print(f"创建Excel失败: {str(e)}")
            return None

    def _auto_adjust_column_width(self, worksheet, df):
        """自动调整列宽"""
        column_max_length = {}
        # 计算表头长度
        for col_idx, col_name in enumerate(df.columns, 1):
            column_max_length[col_idx] = len(str(col_name)) + 3
        # 计算数据长度
        for col_idx, col_name in enumerate(df.columns, 1):
            for value in df[col_name]:
                value_length = len(str(value)) + 2
                if value_length > column_max_length[col_idx]:
                    column_max_length[col_idx] = value_length

        # 设置列宽（限制最大宽度）
        max_width = 22
        for col_idx, max_length in column_max_length.items():
            col_width = min(max_length * 0.9, max_width)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = col_width

    def _add_cell_borders(self, worksheet):
        """添加单元格边框"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in worksheet.iter_rows(
            min_row=1, max_row=worksheet.max_row,
            min_col=1, max_col=worksheet.max_column
        ):
            for cell in row:
                cell.border = thin_border

    def get_available_font(self):
        """获取可用中文字体"""
        font_candidates = [
            'C:/Windows/Fonts/simhei.ttf',
            'C:/Windows/Fonts/msyh.ttf',
            'C:/Windows/Fonts/simsun.ttc',
        ]
        for font_path in font_candidates:
            if os.path.exists(font_path):
                return font_path
        return None

    def _excel_to_pdf(self, excel_path, pdf_path, max_rows=300):
        """Excel转PDF（纵向分页）"""
        if not os.path.exists(excel_path):
            print(f"Excel不存在: {excel_path}")
            return False

        try:
            # Windows系统用Excel COM（最佳效果）
            if os.name == 'nt':
                return self._excel_to_pdf_via_com(excel_path, pdf_path, max_rows)
            # 其他系统用FPDF模拟
            return self._excel_to_pdf_via_fpdf(excel_path, pdf_path, max_rows)
        except Exception as e:
            print(f"PDF转换失败: {str(e)}")
            return False

    def _excel_to_pdf_via_com(self, excel_path, pdf_path, max_rows):
        """Windows专用：Excel COM转换PDF"""
        pythoncom.CoInitialize()
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            workbook = excel.Workbooks.Open(
                Filename=os.path.abspath(excel_path),
                ReadOnly=True
            )
            worksheet = workbook.ActiveSheet

            # 限制最大行数
            total_rows = worksheet.UsedRange.Rows.Count
            if total_rows > max_rows:
                for row in range(max_rows + 1, total_rows + 1):
                    worksheet.Rows(row).Hidden = True

            # 关键设置：纵向分页
            worksheet.PageSetup.Orientation = 1  # 纵向
            worksheet.PageSetup.FitToPagesWide = 1  # 宽度适配1页
            worksheet.PageSetup.PrintTitleRows = "$1:$1"  # 每页显示表头

            # 导出PDF
            worksheet.ExportAsFixedFormat(
                Type=0,
                Filename=os.path.abspath(pdf_path)
            )

            workbook.Close(False)
            excel.Quit()
            print(f"PDF生成成功: {pdf_path}")
            return True
        except Exception as e:
            print(f"COM转换出错: {str(e)}")
            return False
        finally:
            pythoncom.CoUninitialize()

    def _excel_to_pdf_via_fpdf(self, excel_path, pdf_path, max_rows):
        """非Windows系统：FPDF转换PDF"""
        try:
            df = pd.read_excel(excel_path)
            if len(df) > max_rows:
                df = df.head(max_rows)

            # A4纵向
            pdf = FPDF(orientation='P', unit='mm', format='A4')
            pdf.add_page()

            # 设置字体
            font_path = self.get_available_font()
            font_name = 'Arial'
            if font_path:
                try:
                    pdf.add_font('Chinese', '', font_path, uni=True)
                    pdf.set_font('Chinese', size=9)
                    font_name = 'Chinese'
                except:
                    pdf.set_font('Arial', size=9)
            else:
                pdf.set_font('Arial', size=9)

            # 计算列宽
            col_width = 25
            # 表头
            pdf.set_fill_color(240, 240, 240)
            for col_name in df.columns:
                pdf.cell(col_width, 10, str(col_name), border=1, fill=True)
            pdf.ln()

            # 数据行
            pdf.set_fill_color(255, 255, 255)
            for _, row in df.iterrows():
                if pdf.get_y() > 270:  # 换页检查
                    pdf.add_page()
                    # 重新打印表头
                    pdf.set_fill_color(240, 240, 240)
                    for col_name in df.columns:
                        pdf.cell(col_width, 10, str(col_name), border=1, fill=True)
                    pdf.ln()
                    pdf.set_fill_color(255, 255, 255)
                
                for col_name in df.columns:
                    value = str(row[col_name]) if pd.notna(row[col_name]) else ""
                    # 截断过长的值
                    if len(value) > 20:
                        value = value[:20] + "..."
                    pdf.cell(col_width, 10, value, border=1)
                pdf.ln()

            pdf.output(pdf_path)
            print(f"PDF生成成功: {pdf_path}")
            return True
        except Exception as e:
            print(f"FPDF转换出错: {str(e)}")
            return False

    def _analyze_and_generate_sql(self, query_results, table_name):
        """分析数据并生成清洗SQL"""
        print("===== 分析数据质量 =====")
        
        analysis_prompt = f"""
        基于以下查询结果，分析数据质量并生成清洗SQL：

        表名: {table_name}
        
        基本信息：{query_results.get('基本信息', '无')}
        缺失值：{query_results.get('缺失值分析', '无')}
        重复数据：{query_results.get('重复数据分析', '无')}
        数值列：{query_results.get('数值列分析', '无')}
        文本列：{query_results.get('文本列分析', '无')}

        请提供：
        1. 数据集概述
        2. 质量问题
        3. 清洗建议
        4. 清洗SQL（前缀"EXECUTE SQL:"）
        """

        messages = [
            SystemMessage(content="专业数据分析师，擅长SQL清洗"),
            HumanMessage(content=analysis_prompt)
        ]

        response = self.llm.invoke(messages)
        return response.content

    def _cleanup_connections(self):
        """清理所有数据库连接"""
        if self.raw_connection:
            try:
                self.raw_connection.close()
            except Exception as e:
                print(f"关闭数据库连接出错: {str(e)}")
        self.raw_connection = None
        self.current_table_name = None

    def process_dataset(self, file_path):
        """主流程"""
        temp_excel = None
        start_time = time.time()
        try:
            # 创建内存数据库
            db_result = self.create_in_memory_db(file_path)
            if not isinstance(db_result, tuple) or len(db_result) < 2:
                return db_result[1] if isinstance(db_result, tuple) else str(db_result)

            df, table_name = db_result

            print("===== 开始执行数据查询 =====")

            # 同步执行查询
            query_results = {}
            
            print("===== 执行查询：基本信息 =====")
            query_results['基本信息'] = self.get_basic_info(table_name)
            
            print("===== 执行查询：缺失值分析 =====")
            query_results['缺失值分析'] = self.get_missing_value_analysis(table_name)
            
            print("===== 执行查询：重复数据分析 =====")
            query_results['重复数据分析'] = self.get_duplicate_analysis(table_name)
            
            print("===== 执行查询：数值列分析 =====")
            query_results['数值列分析'] = self.get_numeric_analysis(table_name)
            
            print("===== 执行查询：文本列分析 =====")
            query_results['文本列分析'] = self.get_text_analysis(table_name)

            # 分析并生成清洗SQL
            analysis_result = self._analyze_and_generate_sql(query_results, table_name)

            # 获取最终数据
            final_df = pd.read_sql(f"SELECT * FROM {table_name}", self.raw_connection)

            # 生成PDF
            temp_excel = self._create_temp_excel(final_df)
            if not temp_excel:
                return "生成PDF失败：临时Excel创建失败"

            base_name = os.path.splitext(os.path.basename(file_path))[0]
            pdf_path = os.path.join(OUTPUT_PATH, f"{base_name}_样本示例.pdf")

            pdf_success = self._excel_to_pdf(temp_excel, pdf_path)

            # 清理资源
            self._cleanup_connections()
            if temp_excel and os.path.exists(temp_excel):
                os.remove(temp_excel)

            # 计算总运行时间
            end_time = time.time()
            total_time = end_time - start_time
            
            if pdf_success:
                return f"""
===== 处理完成 =====

1. 查询结果:
{chr(10).join(f'{k}:\n{v}\n' for k, v in query_results.items())}

2. 分析与清洗建议:
{analysis_result}

3. 输出文件:
{pdf_path}

4. 运行时间:
总耗时: {total_time:.2f} 秒
"""
        except Exception as e:
            end_time = time.time()
            total_time = end_time - start_time
            
            self._cleanup_connections()
            if temp_excel and os.path.exists(temp_excel):
                try:
                    os.remove(temp_excel)
                except:
                    pass
            return f"处理出错: {str(e)}\n运行时间: {total_time:.2f} 秒"

if __name__ == "__main__":
    processor = DataProcessor()
    file_path = input("请输入数据集文件路径: ")
    
    if os.path.exists(file_path):
        result = processor.process_dataset(file_path)
        print(result)
    else:
        print("文件不存在，请检查路径")
